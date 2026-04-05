"""
REQ-09: Generic Document Retrieval — BM25 full-text search across all worker files.

Indexes domain_data + my_data directories using BM25Okapi (rank_bm25).
In-memory cache with 10-minute TTL; force_refresh=true bypasses cache.

Chunking strategy:
  - Files with <= 2500 words (≈5 pages) → single chunk
  - Larger files → 2500-word chunks

Supported file types: .md .txt .py .rst .yaml .yml .json .csv .docx .pdf .xlsx
"""

import io
import os
import time
import logging
from typing import Optional

from rank_bm25 import BM25Okapi

from sajha.tools.base_mcp_tool import BaseMCPTool
from sajha.core.properties_configurator import PropertiesConfigurator
from sajha.storage import storage

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
WORDS_PER_PAGE = 500
CHUNK_PAGES = 5
CHUNK_WORDS = WORDS_PER_PAGE * CHUNK_PAGES   # 2500
CACHE_TTL = 600                               # seconds

SUPPORTED_EXTENSIONS = {
    '.md', '.txt', '.py', '.rst',
    '.yaml', '.yml', '.json', '.csv',
    '.docx', '.pdf', '.xlsx',
}

# ---------------------------------------------------------------------------
# Module-level cache
# ---------------------------------------------------------------------------
# key  : "<domain_dir>|<my_data_dir>"
# value: {"timestamp": float, "bm25": BM25Okapi|None, "chunks": list}
_INDEX_CACHE: dict = {}


# ---------------------------------------------------------------------------
# Text extraction helpers
# ---------------------------------------------------------------------------

def _extract_text(abs_path: str) -> Optional[str]:
    """Return plain text for a file, or None if unsupported / unreadable."""
    ext = os.path.splitext(abs_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return None
    try:
        raw = storage.read_bytes(abs_path)
    except Exception as exc:
        logger.debug("BM25: cannot read %s: %s", abs_path, exc)
        return None

    try:
        if ext in ('.md', '.txt', '.py', '.rst', '.yaml', '.yml', '.csv'):
            return raw.decode('utf-8', errors='replace')

        if ext == '.json':
            return raw.decode('utf-8', errors='replace')

        if ext == '.docx':
            from docx import Document as _Document
            doc = _Document(io.BytesIO(raw))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

        if ext == '.pdf':
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                pages = [page.extract_text() or '' for page in pdf.pages]
            return '\n'.join(pages)

        if ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cell_text = ' '.join(str(c) for c in row if c is not None)
                    if cell_text.strip():
                        rows.append(cell_text)
            return '\n'.join(rows)

    except Exception as exc:
        logger.debug("BM25: extraction failed for %s: %s", abs_path, exc)

    return None


def _chunk_text(text: str) -> list:
    """Split text into CHUNK_WORDS-word chunks. Small files → single chunk."""
    words = text.split()
    if len(words) <= CHUNK_WORDS:
        return [text]
    return [
        ' '.join(words[i:i + CHUNK_WORDS])
        for i in range(0, len(words), CHUNK_WORDS)
    ]


def _extract_excerpt(text: str, query_tokens: list, context_chars: int = 350) -> str:
    """Return a short excerpt around the first matching query token."""
    lower = text.lower()
    for token in query_tokens:
        pos = lower.find(token)
        if pos >= 0:
            half = context_chars // 2
            start = max(0, pos - half)
            end = min(len(text), pos + len(token) + half)
            excerpt = text[start:end].strip()
            prefix = '...' if start > 0 else ''
            suffix = '...' if end < len(text) else ''
            return prefix + excerpt + suffix
    # No exact token match — return the beginning
    return text[:context_chars].strip() + ('...' if len(text) > context_chars else '')


# ---------------------------------------------------------------------------
# Index builder
# ---------------------------------------------------------------------------

def _build_index(directories: list) -> tuple:
    """
    Scan directories, extract and chunk text, build BM25Okapi index.
    Returns (bm25, chunks_list) — bm25 is None if no indexable content found.
    """
    chunks = []  # list of dicts

    for directory in directories:
        if not storage.exists(directory):
            continue
        rel_paths = storage.list_prefix(directory)
        for rel in rel_paths:
            abs_path = os.path.join(directory, rel)
            text = _extract_text(abs_path)
            if not text or not text.strip():
                continue
            file_chunks = _chunk_text(text.strip())
            n = len(file_chunks)
            for i, chunk_text in enumerate(file_chunks):
                chunks.append({
                    'file_path': abs_path,
                    'file_name': os.path.basename(abs_path),
                    'chunk_index': i,
                    'total_chunks': n,
                    'text': chunk_text,
                })

    if not chunks:
        return None, chunks

    tokenized = [c['text'].lower().split() for c in chunks]
    bm25 = BM25Okapi(tokenized)
    return bm25, chunks


# ---------------------------------------------------------------------------
# Tool
# ---------------------------------------------------------------------------

class DocumentSearchTool(BaseMCPTool):
    """
    BM25 full-text search across domain_data and my_data files.
    Replaces OSFI search and provides generic retrieval for all worker documents.
    """

    def get_input_schema(self) -> dict:
        return self.config.get('inputSchema', {
            'type': 'object',
            'properties': {'query': {'type': 'string'}},
            'required': ['query'],
        })

    def get_output_schema(self) -> dict:
        return {'type': 'object'}

    # ------------------------------------------------------------------
    # Path helpers — mirrors the pattern in operational_tools.py
    # ------------------------------------------------------------------

    def _domain_dir(self) -> str:
        try:
            from flask import g as _g
            root = getattr(_g, 'worker_data_root', None)
            if root:
                return root.rstrip('/')
        except RuntimeError:
            pass
        return PropertiesConfigurator().get(
            'data.domain_data.dir',
            './data/workers/w-market-risk/domain_data',
        )

    def _my_data_dir(self) -> str:
        try:
            from flask import g as _g
            root = getattr(_g, 'worker_my_data_root', None)
            if root:
                return root.rstrip('/')
        except RuntimeError:
            pass
        return PropertiesConfigurator().get(
            'data.my_data.dir',
            './data/workers/w-market-risk/my_data/risk_agent',
        )

    # ------------------------------------------------------------------
    # execute
    # ------------------------------------------------------------------

    def execute(self, arguments: dict) -> dict:
        query = str(arguments.get('query', '')).strip()
        if not query:
            return {'error': 'query is required', 'results': []}

        top_k = max(1, int(arguments.get('top_k', 10)))
        top_n_full = max(0, int(arguments.get('top_n_full_content', 0)))
        force_refresh = bool(arguments.get('force_refresh', False))
        file_type_filter = [
            ft.lower() if ft.startswith('.') else '.' + ft.lower()
            for ft in (arguments.get('file_types') or [])
        ]

        domain_dir = self._domain_dir()
        my_data_dir = self._my_data_dir()
        cache_key = f"{domain_dir}|{my_data_dir}"

        # ── Cache logic ──────────────────────────────────────────────
        cached = _INDEX_CACHE.get(cache_key)
        now = time.time()
        cache_age = round(now - cached['timestamp']) if cached else None
        stale = cached is None or (now - cached['timestamp']) > CACHE_TTL

        if force_refresh or stale:
            bm25, chunks = _build_index([domain_dir, my_data_dir])
            _INDEX_CACHE[cache_key] = {
                'timestamp': now,
                'bm25': bm25,
                'chunks': chunks,
            }
            cache_age = 0
        else:
            bm25 = cached['bm25']
            chunks = cached['chunks']

        if bm25 is None or not chunks:
            return {
                'query': query,
                'total_results': 0,
                'index_size': 0,
                'cache_age_seconds': cache_age,
                'results': [],
                'message': 'No indexable documents found in domain_data or my_data.',
            }

        # ── Apply optional file-type filter ──────────────────────────
        if file_type_filter:
            filtered_idx = [
                i for i, c in enumerate(chunks)
                if os.path.splitext(c['file_name'])[1].lower() in file_type_filter
            ]
        else:
            filtered_idx = list(range(len(chunks)))

        # ── BM25 scoring ─────────────────────────────────────────────
        query_tokens = query.lower().split()
        all_scores = bm25.get_scores(query_tokens)

        scored = sorted(
            ((i, float(all_scores[i])) for i in filtered_idx if all_scores[i] > 0),
            key=lambda x: x[1],
            reverse=True,
        )[:top_k]

        # ── Build results ─────────────────────────────────────────────
        results = []
        for rank, (idx, score) in enumerate(scored):
            chunk = chunks[idx]
            result = {
                'rank': rank + 1,
                'file_name': chunk['file_name'],
                'file_path': chunk['file_path'],
                'chunk_index': chunk['chunk_index'],
                'total_chunks': chunk['total_chunks'],
                'score': round(score, 4),
                'excerpt': _extract_excerpt(chunk['text'], query_tokens),
            }
            if rank < top_n_full:
                result['full_content'] = chunk['text']
            results.append(result)

        return {
            'query': query,
            'total_results': len(results),
            'index_size': len(chunks),
            'cache_age_seconds': cache_age,
            'results': results,
        }
