"""msdoc_get_excel_stats — per-sheet stats for an Excel workbook."""
import io
from typing import Any, Dict

from tools_pack_impl._msdoc_base import MsDocBaseTool, SECTION_PROP


class MsdocGetExcelStats(MsDocBaseTool):
    """Get row/column counts, types, and numeric summary per sheet."""

    def get_input_schema(self) -> Dict:
        return {
            "type": "object",
            "properties": {
                "filename": {"type": "string"},
                "section": SECTION_PROP,
            },
            "required": ["filename"],
        }

    def get_output_schema(self) -> Dict:
        return {"type": "object"}

    def execute(self, arguments: Dict[str, Any]) -> Dict:
        ctx = arguments.get('_worker_context') or {}
        filename = arguments['filename']
        section = arguments.get('section', 'domain_data')
        file_path = self._get_file_path(ctx, filename, section)

        if not self._exists(file_path):
            raise ValueError(f'File not found: {filename}')

        try:
            from openpyxl import load_workbook
            raw = self._read_bytes(file_path)
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            sheets_stats = []

            for sheet in wb.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    sheets_stats.append({
                        'sheet': sheet.title,
                        'row_count': 0,
                        'column_count': 0,
                        'columns': [],
                        'numeric_summary': {},
                    })
                    continue

                header = [str(c) if c is not None else f'col_{i}' for i, c in enumerate(rows[0])]
                data_rows = rows[1:]
                col_count = len(header)

                col_stats = {}
                for ci, col_name in enumerate(header):
                    values = [r[ci] for r in data_rows if ci < len(r) and r[ci] is not None]
                    nums = []
                    for v in values:
                        try:
                            nums.append(float(v))
                        except (TypeError, ValueError):
                            pass
                    if nums:
                        col_stats[col_name] = {
                            'type': 'numeric',
                            'non_null': len(values),
                            'min': round(min(nums), 4),
                            'max': round(max(nums), 4),
                            'mean': round(sum(nums) / len(nums), 4),
                        }
                    else:
                        col_stats[col_name] = {
                            'type': 'text',
                            'non_null': len(values),
                        }

                sheets_stats.append({
                    'sheet': sheet.title,
                    'row_count': len(data_rows),
                    'column_count': col_count,
                    'columns': header,
                    'column_stats': col_stats,
                })

            wb.close()
            return {
                'filename': filename,
                'sheet_count': len(sheets_stats),
                'sheets': sheets_stats,
                '_source': file_path,
            }
        except ImportError:
            raise ValueError('openpyxl library not installed. Install with: pip install openpyxl')
        except Exception as e:
            raise ValueError(f'Failed to get Excel stats: {str(e)}')
