"""
BM25 document search — compliant with upstream SAJHA v5.0.0.

Ported from sajhamcpserver/sajha/tools/impl/bm25_search_tool.py (our pre-migration fork).

Compliance changes from the original:
- Extends upstream's `sajha.tools.base_mcp_tool.BaseMCPTool` (provides `execute_with_tracking`).
- Worker context comes from `arguments['_worker_context']` (a dict injected by our agent
  in agent/tools.py:_service_headers/_call_sajha), not from Flask `g`.
- Uses `pathlib`/`os` directly for filesystem access — no dependency on the old fork's
  `sajha.storage` abstraction.
- No dependency on the old `PropertiesConfigurator`.
"""

import io
import os
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from rank_bm25 import BM25Okapi
from sajha.tools.base_mcp_tool import BaseMCPTool

from tools_pack_lib.worker_ctx import get_data_layers

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    '.md', '.txt', '.py', '.rst', '.yaml', '.yml', '.json', '.csv',
    '.docx', '.pdf', '.xlsx',
}

# Module-level cache: key = "|".join(layer roots), value = {fingerprint, bm25, docs}
_INDEX_CACHE: Dict[str, Dict[str, Any]] = {}


def _fingerprint(directories: List[str]) -> Dict[str, int]:
    fp: Dict[str, int] = {}
    for d in directories:
        if not d or not os.path.isdir(d):
            continue
        for root, _, files in os.walk(d):
            for f in files:
                p = os.path.join(root, f)
                try:
                    fp[p] = os.path.getsize(p)
                except OSError:
                    pass
    return fp


def _extract_text(abs_path: str) -> Optional[str]:
    ext = os.path.splitext(abs_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return None
    try:
        with open(abs_path, 'rb') as fh:
            raw = fh.read()
    except OSError as exc:
        logger.debug("BM25: cannot read %s: %s", abs_path, exc)
        return None

    try:
        if ext in ('.md', '.txt', '.py', '.rst', '.yaml', '.yml', '.csv', '.json'):
            return raw.decode('utf-8', errors='replace')
        if ext == '.docx':
            from docx import Document as _Document
            doc = _Document(io.BytesIO(raw))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
        if ext == '.pdf':
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                return '\n'.join((page.extract_text() or '') for page in pdf.pages)
        if ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            chunks: List[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        chunks.append(' '.join(cells))
            return '\n'.join(chunks)
    except Exception as exc:
        logger.debug("BM25: extraction failed for %s: %s", abs_path, exc)
    return None


def _build_index(directories: List[str]):
    docs: List[Dict[str, Any]] = []
    for d in directories:
        if not d or not os.path.isdir(d):
            continue
        for root, _, files in os.walk(d):
            for f in files:
                if f.startswith('.') or f.startswith('_'):
                    continue
                abs_path = os.path.join(root, f)
                text = _extract_text(abs_path)
                if not text:
                    continue
                docs.append({
                    'file_name': f,
                    'file_path': abs_path,
                    'text': text,
                })
    if not docs:
        return None, []
    tokenised = [d['text'].lower().split() for d in docs]
    bm25 = BM25Okapi(tokenised)
    return bm25, docs


def _extract_excerpt(text: str, query_tokens: List[str], window: int = 240) -> str:
    if not text:
        return ''
    lower = text.lower()
    for tok in query_tokens:
        idx = lower.find(tok)
        if idx >= 0:
            start = max(0, idx - window // 2)
            end = min(len(text), idx + window // 2)
            return ('…' if start > 0 else '') + text[start:end].strip() + ('…' if end < len(text) else '')
    return text[:window].strip() + ('…' if len(text) > window else '')


class BM25SearchTool(BaseMCPTool):
    """REQ-09 BM25 document search across worker data layers.

    Worker context flows in via `arguments['_worker_context']` — see
    tools_pack.lib.worker_ctx for the header→ctx mapping.
    """

    def get_input_schema(self) -> Dict:
        return self.config.get('inputSchema', {
            'type': 'object',
            'properties': {
                'query':           {'type': 'string', 'description': 'Free-text search query'},
                'top_k':           {'type': 'integer', 'description': 'Max results', 'default': 10},
                'file_types':      {'type': 'array', 'items': {'type': 'string'}, 'description': 'Filter by extension'},
                'top_n_full_content': {'type': 'integer', 'description': 'Include full content for top N', 'default': 0},
            },
            'required': ['query'],
        })

    def get_output_schema(self) -> Dict:
        return {
            'type': 'object',
            'properties': {
                'query':         {'type': 'string'},
                'total_results': {'type': 'integer'},
                'index_size':    {'type': 'integer'},
                'rebuilt':       {'type': 'boolean'},
                'results':       {'type': 'array'},
            },
        }

    def execute(self, arguments: Dict[str, Any]) -> Dict[str, Any]:
        query = str(arguments.get('query', '')).strip()
        if not query:
            return {'error': 'query is required', 'results': []}

        top_k = max(1, int(arguments.get('top_k', 10)))
        top_n_full = max(0, int(arguments.get('top_n_full_content', 0)))
        file_type_filter = [
            ft.lower() if ft.startswith('.') else '.' + ft.lower()
            for ft in (arguments.get('file_types') or [])
        ]

        ctx = arguments.get('_worker_context') or {}
        layers = get_data_layers(ctx, 'all')
        directories = [p for _, p in layers]
        cache_key = '|'.join(directories) or '<empty>'

        # Fingerprint freshness check
        current_fp = _fingerprint(directories)
        cached = _INDEX_CACHE.get(cache_key)
        rebuilt = False
        if cached is None or cached['fingerprint'] != current_fp:
            bm25, docs = _build_index(directories)
            _INDEX_CACHE[cache_key] = {'fingerprint': current_fp, 'bm25': bm25, 'docs': docs}
            rebuilt = True
        else:
            bm25, docs = cached['bm25'], cached['docs']

        if bm25 is None or not docs:
            return {
                'query': query, 'total_results': 0, 'index_size': 0, 'rebuilt': rebuilt,
                'results': [], 'message': 'No indexable documents found in worker data layers.',
                '_searched_dirs': directories,
            }

        if file_type_filter:
            filtered_idx = [i for i, d in enumerate(docs)
                            if os.path.splitext(d['file_name'])[1].lower() in file_type_filter]
        else:
            filtered_idx = list(range(len(docs)))

        query_tokens = query.lower().split()
        all_scores = bm25.get_scores(query_tokens)

        scored = sorted(
            ((i, float(all_scores[i])) for i in filtered_idx if all_scores[i] > 0),
            key=lambda x: x[1], reverse=True,
        )[:top_k]

        results: List[Dict[str, Any]] = []
        for rank, (idx, score) in enumerate(scored):
            doc = docs[idx]
            entry = {
                'rank': rank + 1,
                'file_name': doc['file_name'],
                'file_path': doc['file_path'],
                'score': round(score, 4),
                'excerpt': _extract_excerpt(doc['text'], query_tokens),
            }
            if rank < top_n_full:
                entry['full_content'] = doc['text']
            results.append(entry)

        return {
            'query': query,
            'total_results': len(results),
            'index_size': len(docs),
            'rebuilt': rebuilt,
            'results': results,
        }
